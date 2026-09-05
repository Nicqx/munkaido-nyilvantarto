from __future__ import annotations

import json
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.datetime import from_excel

from .core import DAY_TYPE_LABELS, parse_hms


SEED_META_KEY = "seed_excel_import_v1"
KNOWN_BAD_DATE = date(2026, 7, 27)
IMPORT_FORMAT_VERSION = 1
DETAIL_SHEET = "Munkaidőadatok"
ANNUAL_SHEET = "Éves összesítő"
MAX_IMPORT_ROWS = 5000

DETAIL_HEADERS = (
    "Dátum",
    "Nap",
    "Típus",
    "Érkezés",
    "Távozás",
    "Kint töltött idő",
    "Egyedi elvárt idő",
    "Megjegyzés",
    "Ledolgozott idő",
    "Elvárt idő",
    "Egyenleg",
    "Egyenleg (óra)",
)

# A 1.2.0-s kiadás exportja is visszatölthető. Abban még nem szerepelt
# külön oszlopban az egyedi elvárt idő.
LEGACY_DETAIL_HEADERS = (
    "Dátum",
    "Nap",
    "Típus",
    "Érkezés",
    "Távozás",
    "Kint töltött idő",
    "Ledolgozott idő",
    "Elvárt idő",
    "Egyenleg",
    "Egyenleg (óra)",
    "Megjegyzés",
)

ANNUAL_HEADERS = (
    "Év",
    "Éves egyenleg",
    "Egyenleg (óra)",
    "Szabadságkeret",
    "Felhasznált",
    "Fennmaradó",
)


class WorkbookImportError(ValueError):
    def __init__(self, errors: list[str]):
        super().__init__(errors[0] if errors else "Az Excel-fájl nem importálható.")
        self.errors = errors


@dataclass(frozen=True)
class ImportedEntry:
    work_date: date
    arrival_time: str | None
    departure_time: str | None
    break_seconds: int
    day_type: str
    expected_seconds_override: int | None
    note: str


def as_date(value, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    return None


def as_seconds(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, datetime):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, timedelta):
        return max(0, int(round(value.total_seconds())))
    if isinstance(value, (int, float)):
        return max(0, int(round(float(value) * 86400)))
    return None


def import_date(value, epoch) -> date:
    converted = as_date(value, epoch)
    if converted is None and isinstance(value, str):
        try:
            converted = date.fromisoformat(value.strip())
        except ValueError:
            converted = None
    if converted is None or not 1900 <= converted.year <= 9999:
        raise ValueError("a dátum formátuma ÉÉÉÉ-HH-NN legyen")
    return converted


def import_seconds(value, *, allow_over_24: bool = False) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, timedelta) and value.total_seconds() < 0:
        raise ValueError("az idő nem lehet negatív")
    if isinstance(value, (int, float)) and value < 0:
        raise ValueError("az idő nem lehet negatív")
    converted = as_seconds(value)
    if converted is None and isinstance(value, str):
        converted = parse_hms(value, allow_over_24=allow_over_24)
    if converted is None:
        raise ValueError("az idő formátuma ÓÓ:PP vagy ÓÓ:PP:MM legyen")
    if not allow_over_24 and converted >= 24 * 3600:
        raise ValueError("az időpont nem lehet 24:00:00 vagy későbbi")
    if allow_over_24 and converted >= 1000 * 3600:
        raise ValueError("az időtartam legfeljebb 999:59:59 lehet")
    return converted


def header_values(sheet) -> tuple[str, ...]:
    values = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]
    while values and not values[-1]:
        values.pop()
    return tuple(values)


def day_type_from_value(value) -> str:
    text = str(value or "").strip()
    if text in DAY_TYPE_LABELS:
        return text
    labels = {label.casefold(): code for code, label in DAY_TYPE_LABELS.items()}
    day_type = labels.get(text.casefold())
    if not day_type:
        allowed = ", ".join(DAY_TYPE_LABELS.values())
        raise ValueError(f"ismeretlen naptípus; választható: {allowed}")
    return day_type


def parse_detail_sheet(workbook) -> list[ImportedEntry]:
    if DETAIL_SHEET not in workbook.sheetnames:
        raise WorkbookImportError([f'Hiányzik a „{DETAIL_SHEET}” munkalap.'])
    sheet = workbook[DETAIL_SHEET]
    headers = header_values(sheet)
    if headers == DETAIL_HEADERS:
        legacy = False
    elif headers == LEGACY_DETAIL_HEADERS:
        legacy = True
    else:
        raise WorkbookImportError([
            "A Munkaidőadatok fejléc nem megfelelő. Használd az alkalmazásból letöltött Excel-exportot."
        ])
    if sheet.max_row - 1 > MAX_IMPORT_ROWS:
        raise WorkbookImportError([
            f"Legfeljebb {MAX_IMPORT_ROWS} munkanap tölthető fel egyszerre."
        ])

    entries: list[ImportedEntry] = []
    dates: dict[date, int] = {}
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            work_date = import_date(row[0] if len(row) > 0 else None, workbook.epoch)
            if work_date in dates:
                raise ValueError(f"a dátum már szerepel a(z) {dates[work_date]}. sorban")
            day_type = day_type_from_value(row[2] if len(row) > 2 else None)
            arrival_seconds = import_seconds(row[3] if len(row) > 3 else None)
            departure_seconds = import_seconds(row[4] if len(row) > 4 else None)
            break_seconds = import_seconds(
                row[5] if len(row) > 5 else None,
                allow_over_24=True,
            ) or 0

            if legacy:
                effective_expected = import_seconds(
                    row[7] if len(row) > 7 else None,
                    allow_over_24=True,
                )
                expected_override = (
                    effective_expected * 2
                    if effective_expected is not None and day_type.startswith("leave_half")
                    else effective_expected
                )
                note_value = row[10] if len(row) > 10 else None
            else:
                expected_override = import_seconds(
                    row[6] if len(row) > 6 else None,
                    allow_over_24=True,
                )
                note_value = row[7] if len(row) > 7 else None

            if arrival_seconds is not None and departure_seconds is not None:
                if departure_seconds < arrival_seconds:
                    raise ValueError("a távozás nem lehet korábbi az érkezésnél")
                if break_seconds > departure_seconds - arrival_seconds:
                    raise ValueError("a kint töltött idő hosszabb a bent töltött időnél")
            if day_type in ("leave_full", "off"):
                if arrival_seconds is not None or departure_seconds is not None or break_seconds:
                    raise ValueError("szabadságon vagy nem számolt napon nem lehet időadat")
                if expected_override is not None:
                    raise ValueError("szabadságon vagy nem számolt napon nem lehet egyedi elvárt idő")

            note = "" if note_value is None else str(note_value).strip()
            entries.append(ImportedEntry(
                work_date=work_date,
                arrival_time=seconds_to_clock(arrival_seconds),
                departure_time=seconds_to_clock(departure_seconds),
                break_seconds=break_seconds,
                day_type=day_type,
                expected_seconds_override=expected_override,
                note=note,
            ))
            dates[work_date] = row_number
        except ValueError as exc:
            errors.append(f"{row_number}. sor: {exc}.")

    if errors:
        raise WorkbookImportError(errors)
    return entries


def parse_leave_allowances(workbook) -> dict[int, int]:
    if ANNUAL_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[ANNUAL_SHEET]
    if header_values(sheet) != ANNUAL_HEADERS:
        return {}

    allowances: dict[int, int] = {}
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        raw_allowance = row[3] if len(row) > 3 else None
        if raw_allowance in (None, "", "Nincs megadva"):
            continue
        try:
            raw_year = row[0]
            year = int(raw_year)
            if isinstance(raw_year, float) and raw_year != year:
                raise ValueError
            allowance = float(str(raw_allowance).replace(",", "."))
            half_days = round(allowance * 2)
            if not 1900 <= year <= 9999 or allowance < 0 or abs(allowance * 2 - half_days) > 0.000001:
                raise ValueError
            if year in allowances:
                raise ValueError
            allowances[year] = half_days
        except (OverflowError, TypeError, ValueError):
            errors.append(
                f"Az Éves összesítő {row_number}. sorában hibás az év vagy a szabadságkeret."
            )
    if errors:
        raise WorkbookImportError(errors)
    return allowances


def import_export_workbook(
    conn: sqlite3.Connection,
    user_id: int,
    workbook_file,
    *,
    overwrite: bool = False,
) -> dict[str, int]:
    try:
        workbook = load_workbook(workbook_file, data_only=False, read_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise WorkbookImportError([
            "A feltöltött fájl nem olvasható .xlsx munkafüzet."
        ]) from exc

    try:
        entries = parse_detail_sheet(workbook)
        allowances = parse_leave_allowances(workbook)
    finally:
        workbook.close()

    now = datetime.now(timezone.utc).isoformat()
    result = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "allowances_created": 0,
        "allowances_updated": 0,
        "allowances_skipped": 0,
    }
    conn.execute("SAVEPOINT excel_import")
    try:
        for entry in entries:
            existing = conn.execute(
                "SELECT id FROM work_entries WHERE user_id = ? AND work_date = ?",
                (user_id, entry.work_date.isoformat()),
            ).fetchone()
            values = (
                entry.arrival_time,
                entry.departure_time,
                entry.break_seconds,
                entry.day_type,
                entry.expected_seconds_override,
                entry.note,
                now,
            )
            if existing and not overwrite:
                result["skipped"] += 1
            elif existing:
                conn.execute(
                    """
                    UPDATE work_entries
                    SET arrival_time = ?, departure_time = ?, break_seconds = ?,
                        break_started_at = NULL, day_type = ?,
                        expected_seconds_override = ?, note = ?,
                        source = 'excel-import', updated_at = ?
                    WHERE id = ?
                    """,
                    (*values, existing["id"]),
                )
                result["updated"] += 1
            else:
                conn.execute(
                    """
                    INSERT INTO work_entries(
                        user_id, work_date, arrival_time, departure_time,
                        break_seconds, break_started_at, day_type,
                        expected_seconds_override, note, source, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, NULL, ?, ?, ?, 'excel-import', ?, ?)
                    """,
                    (
                        user_id,
                        entry.work_date.isoformat(),
                        entry.arrival_time,
                        entry.departure_time,
                        entry.break_seconds,
                        entry.day_type,
                        entry.expected_seconds_override,
                        entry.note,
                        now,
                        now,
                    ),
                )
                result["created"] += 1

        for year, half_days in allowances.items():
            existing = conn.execute(
                "SELECT 1 FROM leave_allowances WHERE user_id = ? AND year = ?",
                (user_id, year),
            ).fetchone()
            if existing and not overwrite:
                result["allowances_skipped"] += 1
            elif existing:
                conn.execute(
                    "UPDATE leave_allowances SET allowance_half_days = ? WHERE user_id = ? AND year = ?",
                    (half_days, user_id, year),
                )
                result["allowances_updated"] += 1
            else:
                conn.execute(
                    "INSERT INTO leave_allowances(user_id, year, allowance_half_days) VALUES (?, ?, ?)",
                    (user_id, year, half_days),
                )
                result["allowances_created"] += 1
        conn.execute("RELEASE SAVEPOINT excel_import")
    except Exception:
        conn.execute("ROLLBACK TO SAVEPOINT excel_import")
        conn.execute("RELEASE SAVEPOINT excel_import")
        raise
    return result


def seconds_to_clock(value: int | None) -> str | None:
    if value is None:
        return None
    return f"{value // 3600:02d}:{value % 3600 // 60:02d}:{value % 60:02d}"


def combined_note(name, column_k, column_l) -> str:
    parts = []
    for item in (column_k, column_l):
        text = str(item).strip() if item is not None else ""
        if text and text not in parts:
            parts.append(text)
    if name and "MNAP" in str(name).upper():
        parts.insert(0, str(name).strip())
    return " · ".join(parts)


def classify_day(name, note: str) -> str:
    folded = note.casefold()
    if "szabadság" in folded:
        return "leave_full"
    off_markers = ("húsvét", "ünnep", "pünkösd", "pn")
    if any(marker in folded for marker in off_markers) or "MNAP" in str(name or "").upper():
        return "off"
    return "work"


def import_seed_workbook(conn: sqlite3.Connection, user_id: int, workbook_path: str) -> dict:
    existing = conn.execute("SELECT value FROM app_meta WHERE key = ?", (SEED_META_KEY,)).fetchone()
    if existing:
        return json.loads(existing["value"])
    path = Path(workbook_path)
    if not path.exists():
        result = {"status": "missing", "path": str(path), "imported": 0}
        conn.execute("INSERT INTO app_meta(key, value) VALUES (?, ?)", (SEED_META_KEY, json.dumps(result)))
        return result

    workbook = load_workbook(path, data_only=True, read_only=True)
    imported = 0
    skipped = 0
    now = datetime.now(timezone.utc).isoformat()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=4, values_only=True):
            name = row[0] if len(row) > 0 else None
            work_date = as_date(row[1] if len(row) > 1 else None, workbook.epoch)
            if work_date is None:
                continue
            arrival_seconds = as_seconds(row[2] if len(row) > 2 else None)
            departure_seconds = as_seconds(row[3] if len(row) > 3 else None)
            expected_seconds = as_seconds(row[5] if len(row) > 5 else None)
            note = combined_note(
                name,
                row[10] if len(row) > 10 else None,
                row[11] if len(row) > 11 else None,
            )
            day_type = classify_day(name, note)

            if work_date == KNOWN_BAD_DATE:
                arrival_seconds = None
                departure_seconds = None
                note = (note + " · " if note else "") + "Ellenőrzendő: az eredeti Excel képlete hibás volt"
            if day_type in ("leave_full", "off"):
                arrival_seconds = None
                departure_seconds = None
                expected_override = None
            else:
                expected_override = expected_seconds

            cursor = conn.execute(
                """
                INSERT OR IGNORE INTO work_entries(
                    user_id, work_date, arrival_time, departure_time, break_seconds,
                    break_started_at, day_type, expected_seconds_override, note,
                    source, created_at, updated_at
                ) VALUES (?, ?, ?, ?, 0, NULL, ?, ?, ?, 'excel-import', ?, ?)
                """,
                (
                    user_id,
                    work_date.isoformat(),
                    seconds_to_clock(arrival_seconds),
                    seconds_to_clock(departure_seconds),
                    day_type,
                    expected_override,
                    note,
                    now,
                    now,
                ),
            )
            if cursor.rowcount:
                imported += 1
            else:
                skipped += 1

    workbook.close()
    result = {
        "status": "ok",
        "file": path.name,
        "imported": imported,
        "skipped": skipped,
        "known_bad_date_left_empty": KNOWN_BAD_DATE.isoformat(),
    }
    conn.execute("INSERT INTO app_meta(key, value) VALUES (?, ?)", (SEED_META_KEY, json.dumps(result)))
    return result
