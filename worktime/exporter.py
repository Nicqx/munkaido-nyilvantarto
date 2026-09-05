from __future__ import annotations

import sqlite3
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import DAY_NAMES, DAY_TYPE_LABELS, MONTH_NAMES, entry_metrics, format_duration
from .importer import ANNUAL_HEADERS, DETAIL_HEADERS, IMPORT_FORMAT_VERSION


HEADER_FILL = PatternFill("solid", fgColor="245A73")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEAF1")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def autosize(sheet, maximum: int = 34) -> None:
    for column_cells in sheet.columns:
        width = 0
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 10), maximum)


def style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_export(conn: sqlite3.Connection, user, year: int | None) -> BytesIO:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Munkaidőadatok"
    detail.append(list(DETAIL_HEADERS))

    params: list = [user["id"]]
    where = "user_id = ?"
    if year is not None:
        where += " AND work_date BETWEEN ? AND ?"
        params.extend([f"{year}-01-01", f"{year}-12-31"])
    entries = conn.execute(
        f"SELECT * FROM work_entries WHERE {where} ORDER BY work_date",
        params,
    ).fetchall()

    monthly: dict[tuple[int, int], dict[str, int]] = {}
    for entry in entries:
        day = date.fromisoformat(entry["work_date"])
        metrics = entry_metrics(conn, user["id"], day, entry)
        key = (day.year, day.month)
        bucket = monthly.setdefault(key, {"worked": 0, "expected": 0, "balance": 0, "complete": 0})
        if metrics.included:
            bucket["worked"] += int(metrics.worked_seconds or 0)
            bucket["expected"] += int(metrics.expected_seconds or 0)
            bucket["balance"] += int(metrics.balance_seconds or 0)
            bucket["complete"] += 1
        detail.append([
            day,
            DAY_NAMES[day.weekday()],
            DAY_TYPE_LABELS.get(entry["day_type"], entry["day_type"]),
            entry["arrival_time"] or "",
            entry["departure_time"] or "",
            format_duration(entry["break_seconds"] or 0),
            (
                format_duration(entry["expected_seconds_override"])
                if entry["expected_seconds_override"] is not None
                else ""
            ),
            entry["note"],
            format_duration(metrics.worked_seconds) if metrics.included else "",
            format_duration(metrics.expected_seconds) if metrics.included else "",
            format_duration(metrics.balance_seconds, signed=True) if metrics.included else "",
            round((metrics.balance_seconds or 0) / 3600, 4) if metrics.included else None,
        ])
        detail.cell(detail.max_row, 1).number_format = "yyyy-mm-dd"
        note_cell = detail.cell(detail.max_row, 8)
        note_cell.data_type = "s"
        balance_cell = detail.cell(detail.max_row, 11)
        if metrics.included:
            balance_cell.font = Font(color="16803A" if (metrics.balance_seconds or 0) >= 0 else "B42318")

    style_header(detail)
    for cell in detail[1][8:]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(color="245A73", bold=True)
    autosize(detail)

    summary = workbook.create_sheet("Havi összesítő")
    summary.append(["Év", "Hónap", "Ledolgozott", "Elvárt", "Havi egyenleg", "Egyenleg (óra)", "Lezárt napok"])
    cumulative_by_year: dict[int, int] = {}
    for (row_year, month), values in sorted(monthly.items()):
        cumulative_by_year[row_year] = cumulative_by_year.get(row_year, 0) + values["balance"]
        summary.append([
            row_year,
            MONTH_NAMES[month],
            format_duration(values["worked"]),
            format_duration(values["expected"]),
            format_duration(values["balance"], signed=True),
            round(values["balance"] / 3600, 4),
            values["complete"],
        ])
    style_header(summary)
    autosize(summary)

    annual = workbook.create_sheet("Éves összesítő")
    annual.append(list(ANNUAL_HEADERS))
    allowance_years = {
        int(row["year"])
        for row in conn.execute(
            "SELECT year FROM leave_allowances WHERE user_id = ?",
            (user["id"],),
        ).fetchall()
        if year is None or int(row["year"]) == year
    }
    years = sorted(
        {key[0] for key in monthly}
        | allowance_years
        | ({year} if year else set())
    )
    for row_year in years:
        allowance = conn.execute(
            "SELECT allowance_half_days FROM leave_allowances WHERE user_id = ? AND year = ?",
            (user["id"], row_year),
        ).fetchone()
        used = conn.execute(
            """
            SELECT COALESCE(SUM(CASE day_type WHEN 'leave_full' THEN 2
                WHEN 'leave_half_am' THEN 1 WHEN 'leave_half_pm' THEN 1 ELSE 0 END), 0) AS units
            FROM work_entries WHERE user_id = ? AND work_date BETWEEN ? AND ?
            """,
            (user["id"], f"{row_year}-01-01", f"{row_year}-12-31"),
        ).fetchone()["units"]
        total = allowance["allowance_half_days"] if allowance else None
        annual.append([
            row_year,
            format_duration(cumulative_by_year.get(row_year, 0), signed=True),
            round(cumulative_by_year.get(row_year, 0) / 3600, 4),
            total / 2 if total is not None else "Nincs megadva",
            used / 2,
            (total - used) / 2 if total is not None else "Nincs megadva",
        ])
    style_header(annual)
    autosize(annual)

    info = workbook.create_sheet("Információ")
    info.append(["Munkaidő-nyilvántartó export"])
    info["A1"].fill = HEADER_FILL
    info["A1"].font = WHITE_FONT
    info.append(["Felhasználó", user["login"]])
    info["B2"].data_type = "s"
    info.append(["Időszak", str(year) if year else "Minden adat"])
    info.append(["Importformátum verzió", IMPORT_FORMAT_VERSION])
    info.append(["Számítás", "Ledolgozott = távozás − érkezés − kint töltött idő"])
    info.append(["Egyenleg", "Ledolgozott idő − elvárt idő"])
    info.append([
        "Visszatöltés",
        "A Munkaidőadatok első 8 oszlopa és az Éves összesítő szabadságkerete importálható.",
    ])
    autosize(info, 70)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
