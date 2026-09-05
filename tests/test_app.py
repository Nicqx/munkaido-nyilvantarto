import os
import sqlite3
import tempfile
import unittest
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from worktime import create_app
from worktime.core import entry_metrics
from worktime.importer import DETAIL_HEADERS, LEGACY_DETAIL_HEADERS


class AppTests(unittest.TestCase):
    ADMIN_PASSWORD = "test-admin-password"
    SEED_LOGIN = "seed.user@example.test"
    SEED_PASSWORD = "test-seed-password"
    DEFAULT_USER_PASSWORD = "test-reset-password"

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "test.db")
        self.app = create_app({
            "TESTING": True,
            "SECRET_KEY": "test-secret",
            "DATABASE": self.db_path,
            "IMPORT_SEED_EXCEL": False,
            "ADMIN_PASSWORD": self.ADMIN_PASSWORD,
            "DEFAULT_USER_PASSWORD": self.DEFAULT_USER_PASSWORD,
            "SEED_USER_LOGIN": self.SEED_LOGIN,
            "SEED_USER_DISPLAY_NAME": "Teszt importfelhasználó",
            "SEED_USER_PASSWORD": self.SEED_PASSWORD,
        })
        self.client = self.app.test_client()

    def tearDown(self):
        self.tmp.cleanup()

    def login(self, login=None, password=None):
        login = login or self.SEED_LOGIN
        password = password or self.SEED_PASSWORD
        return self.client.post("/login", data={"login": login, "password": password}, follow_redirects=True)

    def test_seed_accounts_can_log_in(self):
        response = self.login()
        self.assertEqual(response.status_code, 200)
        self.assertIn("Napi rögzítés".encode(), response.data)

    def test_registration_creates_user_and_admin_can_see_it(self):
        response = self.client.post(
            "/register",
            data={"display_name": "Új Tesztelő", "email": "uj.tesztelo@example.hu", "password": "Teszt.123"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("A regisztráció elkészült".encode(), response.data)
        login_response = self.login("uj.tesztelo@example.hu", "Teszt.123")
        self.assertIn("Napi rögzítés".encode(), login_response.data)
        self.client.post("/logout")
        self.login("admin", self.ADMIN_PASSWORD)
        admin_page = self.client.get("/admin/users")
        self.assertIn("uj.tesztelo@example.hu".encode(), admin_page.data)

    def test_registration_validation_is_visible_and_preserves_email(self):
        response = self.client.post(
            "/register",
            data={"display_name": "Teszt", "email": "hibas-email", "password": "123"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("A fiók még nem jött létre".encode(), response.data)
        self.assertIn("Adj meg egy használható e-mail-címet".encode(), response.data)
        self.assertIn("A jelszó legalább 4 karakter legyen".encode(), response.data)
        self.assertIn(b'value="hibas-email"', response.data)

    def test_complete_day_and_break_are_calculated(self):
        self.login()
        response = self.client.post(
            "/entry/save",
            data={
                "work_date": "2026-09-07",
                "arrival_time": "08:00:00",
                "departure_time": "18:00:00",
                "break_time": "00:15:00",
                "note": "teszt",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("09:45:00".encode(), response.data)
        self.assertIn("−00:15:00".encode(), response.data)

    def test_full_leave_consumes_one_day(self):
        self.login()
        self.client.post("/leave", data={"year": "2026", "allowance": "25"})
        self.client.post("/entry/leave", data={"work_date": "2026-09-08", "kind": "leave_full"})
        response = self.client.get("/leave?year=2026")
        self.assertIn("24.0 nap".encode(), response.data)

    def test_admin_resets_password(self):
        self.login("admin", self.ADMIN_PASSWORD)
        with self.app.app_context():
            from worktime.db import get_db
            user = get_db().execute(
                "SELECT id FROM users WHERE login = ?", (self.SEED_LOGIN,)
            ).fetchone()
            user_id = user["id"]
        response = self.client.post(f"/admin/users/{user_id}/reset-password", follow_redirects=True)
        self.assertIn("a beállított alapértelmezett jelszóra".encode(), response.data)
        self.client.post("/logout")
        self.assertIn(
            "Napi rögzítés".encode(),
            self.login(self.SEED_LOGIN, self.DEFAULT_USER_PASSWORD).data,
        )

    def test_admin_edits_user_login_and_display_name(self):
        self.client.post(
            "/register",
            data={"display_name": "Régi Név", "email": "regi@example.hu", "password": "Teszt.123"},
        )
        with self.app.app_context():
            from worktime.db import get_db
            user_id = get_db().execute(
                "SELECT id FROM users WHERE login = 'regi@example.hu'"
            ).fetchone()["id"]

        self.login("admin", self.ADMIN_PASSWORD)
        response = self.client.post(
            f"/admin/users/{user_id}/edit",
            data={"display_name": "Új Név", "login": "uj@example.hu"},
            follow_redirects=True,
        )
        self.assertIn("A felhasználó adatai frissültek".encode(), response.data)
        self.assertIn("uj@example.hu".encode(), response.data)
        self.assertIn("Új Név".encode(), response.data)

        self.client.post("/logout")
        login_response = self.login("uj@example.hu", "Teszt.123")
        self.assertIn("Napi rögzítés".encode(), login_response.data)

    def test_admin_rejects_duplicate_login(self):
        self.client.post(
            "/register",
            data={"display_name": "Másik", "email": "masik@example.hu", "password": "Teszt.123"},
        )
        with self.app.app_context():
            from worktime.db import get_db
            user_id = get_db().execute(
                "SELECT id FROM users WHERE login = 'masik@example.hu'"
            ).fetchone()["id"]

        self.login("admin", self.ADMIN_PASSWORD)
        response = self.client.post(
            f"/admin/users/{user_id}/edit",
            data={"display_name": "Másik", "login": "admin"},
        )
        self.assertIn("Ez a felhasználónév már használatban van".encode(), response.data)

    def test_admin_deletes_user_and_all_related_data(self):
        self.client.post(
            "/register",
            data={"display_name": "Törlendő", "email": "torlendo@example.hu", "password": "Teszt.123"},
        )
        self.login("torlendo@example.hu", "Teszt.123")
        self.client.post(
            "/entry/save",
            data={
                "work_date": "2026-09-07",
                "arrival_time": "08:00:00",
                "departure_time": "18:00:00",
                "break_time": "00:00:00",
                "note": "törlési teszt",
            },
        )
        self.client.post("/leave", data={"year": "2026", "allowance": "25"})
        self.client.post("/logout")

        with self.app.app_context():
            from worktime.db import get_db
            user_id = get_db().execute(
                "SELECT id FROM users WHERE login = 'torlendo@example.hu'"
            ).fetchone()["id"]

        self.login("admin", self.ADMIN_PASSWORD)
        response = self.client.post(
            f"/admin/users/{user_id}/delete", follow_redirects=True
        )
        self.assertIn("minden hozzá tartozó adat véglegesen törölve lett".encode(), response.data)
        with self.app.app_context():
            from worktime.db import get_db
            db = get_db()
            self.assertIsNone(db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone())
            for table in ("work_schedules", "work_entries", "leave_allowances"):
                count = db.execute(
                    f"SELECT COUNT(*) FROM {table} WHERE user_id = ?", (user_id,)
                ).fetchone()[0]
                self.assertEqual(count, 0)

    def test_deleted_seed_user_is_not_recreated_on_restart(self):
        self.login("admin", self.ADMIN_PASSWORD)
        with self.app.app_context():
            from worktime.db import get_db
            user_id = get_db().execute(
                "SELECT id FROM users WHERE login = ?", (self.SEED_LOGIN,)
            ).fetchone()["id"]
        self.client.post(f"/admin/users/{user_id}/delete")

        restarted_app = create_app({
            "TESTING": True,
            "SECRET_KEY": "test-secret",
            "DATABASE": self.db_path,
            "IMPORT_SEED_EXCEL": False,
        })
        with restarted_app.app_context():
            from worktime.db import get_db
            self.assertIsNone(
                get_db().execute(
                    "SELECT id FROM users WHERE login = ?", (self.SEED_LOGIN,)
                ).fetchone()
            )

    def test_user_schedule_is_independent_and_controls_balance(self):
        self.client.post(
            "/register",
            data={
                "display_name": "Más Beosztás",
                "email": "mas.beosztas@example.hu",
                "password": "Teszt.123",
            },
        )
        self.login("mas.beosztas@example.hu", "Teszt.123")
        schedule_data = {
            f"{kind}_{weekday}": ""
            for weekday in range(7)
            for kind in ("arrival", "departure")
        }
        schedule_data.update({"arrival_0": "09:00:00", "departure_0": "17:00:00"})
        response = self.client.post("/schedule", data=schedule_data, follow_redirects=True)
        self.assertIn("A saját heti beosztásod elmentve".encode(), response.data)

        self.client.post(
            "/entry/save",
            data={
                "work_date": "2020-01-06",
                "arrival_time": "09:00:00",
                "departure_time": "17:00:00",
                "break_time": "00:00:00",
                "note": "",
            },
        )
        with self.app.app_context():
            from worktime.db import get_db
            db = get_db()
            custom_user = db.execute(
                "SELECT id FROM users WHERE login = 'mas.beosztas@example.hu'"
            ).fetchone()
            seed_user = db.execute(
                "SELECT id FROM users WHERE login = ?", (self.SEED_LOGIN,)
            ).fetchone()
            custom_schedule = db.execute(
                "SELECT expected_seconds FROM work_schedules WHERE user_id = ? AND weekday = 0",
                (custom_user["id"],),
            ).fetchone()
            seed_schedule = db.execute(
                "SELECT expected_seconds FROM work_schedules WHERE user_id = ? AND weekday = 0",
                (seed_user["id"],),
            ).fetchone()
            entry = db.execute(
                "SELECT * FROM work_entries WHERE user_id = ? AND work_date = '2020-01-06'",
                (custom_user["id"],),
            ).fetchone()
            metrics = entry_metrics(db, custom_user["id"], date(2020, 1, 6), entry)
            self.assertEqual(custom_schedule["expected_seconds"], 8 * 3600)
            self.assertEqual(seed_schedule["expected_seconds"], 10 * 3600)
            self.assertEqual(metrics.balance_seconds, 0)

    def test_weekly_fill_uses_defaults_without_overwriting_existing_days(self):
        self.login()
        self.client.post(
            "/entry/save",
            data={
                "work_date": "2020-01-07",
                "arrival_time": "09:15:00",
                "departure_time": "",
                "break_time": "00:00:00",
                "note": "részleges",
            },
        )
        self.client.post(
            "/entry/leave",
            data={"work_date": "2020-01-08", "kind": "leave_full"},
        )
        self.client.post(
            "/entry/save",
            data={
                "work_date": "2020-01-09",
                "arrival_time": "10:00:00",
                "departure_time": "18:00:00",
                "break_time": "00:00:00",
                "note": "meglévő",
            },
        )
        response = self.client.post(
            "/entry/fill-week",
            data={"selected_date": "2020-01-08"},
            follow_redirects=True,
        )
        self.assertIn("Heti pótlás kész: 2 nap kitöltve".encode(), response.data)

        with self.app.app_context():
            from worktime.db import get_db
            rows = get_db().execute(
                """
                SELECT work_date, arrival_time, departure_time, day_type, source
                FROM work_entries
                WHERE user_id = (
                    SELECT id FROM users WHERE login = ?
                ) AND work_date BETWEEN '2020-01-06' AND '2020-01-12'
                ORDER BY work_date
                """,
                (self.SEED_LOGIN,),
            ).fetchall()
            by_day = {row["work_date"]: row for row in rows}
            self.assertEqual(
                (by_day["2020-01-06"]["arrival_time"], by_day["2020-01-06"]["departure_time"]),
                ("08:00:00", "18:00:00"),
            )
            self.assertEqual(by_day["2020-01-06"]["source"], "weekly-auto-fill")
            self.assertEqual(by_day["2020-01-07"]["arrival_time"], "09:15:00")
            self.assertIsNone(by_day["2020-01-07"]["departure_time"])
            self.assertEqual(by_day["2020-01-08"]["day_type"], "leave_full")
            self.assertEqual(by_day["2020-01-09"]["arrival_time"], "10:00:00")
            self.assertEqual(
                (by_day["2020-01-10"]["arrival_time"], by_day["2020-01-10"]["departure_time"]),
                ("08:00:00", "13:30:00"),
            )

    def test_legacy_schedule_schema_is_migrated_without_losing_duration(self):
        legacy_path = os.path.join(self.tmp.name, "legacy.db")
        conn = sqlite3.connect(legacy_path)
        conn.executescript(
            """
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                login TEXT NOT NULL UNIQUE COLLATE NOCASE,
                display_name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );
            CREATE TABLE work_schedules (
                user_id INTEGER NOT NULL,
                weekday INTEGER NOT NULL,
                expected_seconds INTEGER NOT NULL,
                PRIMARY KEY (user_id, weekday)
            );
            INSERT INTO users(
                id, login, display_name, password_hash, is_admin, active, created_at
            ) VALUES (42, 'legacy@example.hu', 'Legacy', 'x', 0, 1, '2020-01-01');
            INSERT INTO work_schedules(user_id, weekday, expected_seconds)
            VALUES (42, 0, 25200);
            """
        )
        conn.close()

        migrated_app = create_app({
            "TESTING": True,
            "SECRET_KEY": "test-secret",
            "DATABASE": legacy_path,
            "IMPORT_SEED_EXCEL": False,
            "ADMIN_PASSWORD": self.ADMIN_PASSWORD,
        })
        with migrated_app.app_context():
            from worktime.db import get_db
            row = get_db().execute(
                """
                SELECT expected_seconds, default_arrival_time, default_departure_time
                FROM work_schedules WHERE user_id = 42 AND weekday = 0
                """
            ).fetchone()
            self.assertEqual(row["expected_seconds"], 7 * 3600)
            self.assertEqual(row["default_arrival_time"], "08:00:00")
            self.assertEqual(row["default_departure_time"], "15:00:00")

    def test_excel_export_is_valid(self):
        self.login()
        response = self.client.get("/export.xlsx?year=2026")
        self.assertEqual(response.status_code, 200)
        target = os.path.join(self.tmp.name, "export.xlsx")
        with open(target, "wb") as handle:
            handle.write(response.data)
        workbook = load_workbook(target, read_only=True)
        self.assertEqual(workbook.sheetnames, ["Munkaidőadatok", "Havi összesítő", "Éves összesítő", "Információ"])
        self.assertEqual(
            tuple(cell.value for cell in workbook["Munkaidőadatok"][1]),
            DETAIL_HEADERS,
        )
        info = {
            row[0].value: row[1].value
            for row in workbook["Információ"].iter_rows(min_row=2, max_col=2)
        }
        self.assertEqual(info["Importformátum verzió"], 1)
        workbook.close()

    def test_export_can_be_imported_back_with_entries_and_leave_allowance(self):
        self.login()
        self.client.post(
            "/entry/save",
            data={
                "work_date": "2026-09-07",
                "arrival_time": "08:00:00",
                "departure_time": "18:00:00",
                "break_time": "00:15:00",
                "note": "=nem Excel-képlet",
            },
        )
        self.client.post(
            "/entry/save",
            data={
                "work_date": "2026-09-08",
                "arrival_time": "12:00:00",
                "departure_time": "16:00:00",
                "break_time": "00:05:00",
                "note": "fél nap",
            },
        )
        self.client.post(
            "/entry/leave",
            data={"work_date": "2026-09-08", "kind": "leave_half_am"},
        )
        self.client.post(
            "/entry/leave",
            data={"work_date": "2026-09-09", "kind": "leave_full"},
        )
        self.client.post("/leave", data={"year": "2026", "allowance": "25,5"})

        exported = self.client.get("/export.xlsx?year=2026").data
        workbook = load_workbook(BytesIO(exported), data_only=False)
        self.assertEqual(workbook["Munkaidőadatok"]["H2"].data_type, "s")
        workbook.close()

        with self.app.app_context():
            from worktime.db import get_db
            db = get_db()
            user_id = db.execute(
                "SELECT id FROM users WHERE login = ?", (self.SEED_LOGIN,)
            ).fetchone()["id"]
            db.execute("DELETE FROM work_entries WHERE user_id = ?", (user_id,))
            db.execute("DELETE FROM leave_allowances WHERE user_id = ?", (user_id,))
            db.commit()

        response = self.client.post(
            "/import",
            data={
                "mode": "skip",
                "file": (BytesIO(exported), "munkaido-2026.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("3 új nap importálva".encode(), response.data)
        self.assertIn("1 szabadságkeret átvéve".encode(), response.data)

        with self.app.app_context():
            from worktime.db import get_db
            db = get_db()
            rows = db.execute(
                """
                SELECT work_date, arrival_time, departure_time, break_seconds,
                       day_type, expected_seconds_override, note
                FROM work_entries WHERE user_id = ? ORDER BY work_date
                """,
                (user_id,),
            ).fetchall()
            self.assertEqual(len(rows), 3)
            self.assertEqual(
                tuple(rows[0]),
                ("2026-09-07", "08:00:00", "18:00:00", 900, "work", None, "=nem Excel-képlet"),
            )
            self.assertEqual(rows[1]["day_type"], "leave_half_am")
            self.assertEqual(rows[1]["break_seconds"], 300)
            self.assertEqual(rows[2]["day_type"], "leave_full")
            allowance = db.execute(
                "SELECT allowance_half_days FROM leave_allowances WHERE user_id = ? AND year = 2026",
                (user_id,),
            ).fetchone()
            self.assertEqual(allowance["allowance_half_days"], 51)

    def test_import_skips_existing_dates_unless_overwrite_is_selected(self):
        self.login()
        original = {
            "work_date": "2026-10-05",
            "arrival_time": "08:00:00",
            "departure_time": "18:00:00",
            "break_time": "00:00:00",
            "note": "exportált érték",
        }
        self.client.post("/entry/save", data=original)
        exported = self.client.get("/export.xlsx?year=2026").data
        changed = dict(original, arrival_time="09:00:00", note="kézzel módosított")
        self.client.post("/entry/save", data=changed)

        response = self.client.post(
            "/import",
            data={"mode": "skip", "file": (BytesIO(exported), "export.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("1 meglévő nap kihagyva".encode(), response.data)

        with self.app.app_context():
            from worktime.db import get_db
            row = get_db().execute(
                "SELECT arrival_time, note FROM work_entries WHERE work_date = '2026-10-05'"
            ).fetchone()
            self.assertEqual(tuple(row), ("09:00:00", "kézzel módosított"))

        response = self.client.post(
            "/import",
            data={"mode": "overwrite", "file": (BytesIO(exported), "export.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("1 meglévő nap felülírva".encode(), response.data)
        with self.app.app_context():
            from worktime.db import get_db
            row = get_db().execute(
                "SELECT arrival_time, note FROM work_entries WHERE work_date = '2026-10-05'"
            ).fetchone()
            self.assertEqual(tuple(row), ("08:00:00", "exportált érték"))

    def test_invalid_import_is_atomic(self):
        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Munkaidőadatok"
        sheet.append(list(DETAIL_HEADERS))
        sheet.append([
            date(2026, 11, 2), "Hétfő", "Munkanap", "08:00:00", "16:00:00",
            "00:00:00", "", "jó sor", "", "", "", "",
        ])
        sheet.append([
            date(2026, 11, 3), "Kedd", "Munkanap", "18:00:00", "08:00:00",
            "00:00:00", "", "hibás sor", "", "", "", "",
        ])
        file_data = BytesIO()
        workbook.save(file_data)
        file_data.seek(0)

        response = self.client.post(
            "/import",
            data={"mode": "skip", "file": (file_data, "hibas.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertIn("az adatbázis nem változott".encode(), response.data)
        self.assertIn("a távozás nem lehet korábbi".encode(), response.data)
        with self.app.app_context():
            from worktime.db import get_db
            count = get_db().execute(
                "SELECT COUNT(*) FROM work_entries WHERE work_date IN ('2026-11-02', '2026-11-03')"
            ).fetchone()[0]
            self.assertEqual(count, 0)

    def test_version_1_2_export_can_also_be_imported(self):
        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Munkaidőadatok"
        sheet.append(list(LEGACY_DETAIL_HEADERS))
        sheet.append([
            date(2026, 12, 1), "Kedd", "Munkanap", "08:00:00", "16:00:00",
            "00:10:00", "07:50:00", "08:00:00", "−00:10:00", -0.1667,
            "régi export",
        ])
        file_data = BytesIO()
        workbook.save(file_data)
        file_data.seek(0)

        response = self.client.post(
            "/import",
            data={"mode": "skip", "file": (file_data, "regi-export.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("1 új nap importálva".encode(), response.data)
        with self.app.app_context():
            from worktime.db import get_db
            row = get_db().execute(
                """
                SELECT break_seconds, expected_seconds_override, note
                FROM work_entries WHERE work_date = '2026-12-01'
                """
            ).fetchone()
            self.assertEqual(tuple(row), (600, 8 * 3600, "régi export"))

    def test_import_requires_login(self):
        response = self.client.get("/import")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])


if __name__ == "__main__":
    unittest.main()
